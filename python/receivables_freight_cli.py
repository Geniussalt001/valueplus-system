import argparse
import csv
import json
import re
import sys
import tempfile
import urllib.request
from pathlib import Path


IV_PATTERN = re.compile(r"^IV(VPR\d+)$", re.IGNORECASE)
DATE_PATTERN = re.compile(r"^\d{1,2}/\d{1,2}/\d{4}$")

WAREHOUSE_RULES = {
    "มหาชัย": ("ซีพี ออลล์ (BDC.มหาชัย)", "สมุทรสาคร"),
    "สำโรง": ("ซีพี ออลล์ (BDC.สำโรง)", "สมุทรปราการ"),
    "ร่มเกล้า": ("ซีพี ออลล์ (BDC.ร่มเกล้า)", "กรุงเทพมหานคร"),
    "ชลบุรี": ("ซีพี ออลล์ (BDC.ชลบุรี)", "กรุงเทพมหานคร"),
    "รังสิต": ("ซีพี ออลล์ (BDC.รังสิต)", "ปทุมธานี"),
    "โชคชัย": ("ซีพี ออลล์ (BDC.โชคชัย)", "กรุงเทพมหานคร"),
    "เชียงใหม่": ("ซีพี ออลล์ (BDC.เชียงใหม่)", "ปทุมธานี"),
    "นครสวรรค์": ("ซีพี ออลล์ (BDC.นครสวรรค์)", "ปทุมธานี"),
    "ขอนแก่น": ("ซีพี ออลล์ (BDC.ขอนแก่น)", "ปทุมธานี"),
    "โคราช": ("ซีพี ออลล์ (BDC.โคราช)", "ปทุมธานี"),
    "หาดใหญ่": ("ซีพี ออลล์ (BDC.หาดใหญ่)", "สมุทรสาคร"),
    "สุราษฎร์ธานี": ("ซีพี ออลล์ (BDC.สุราษฎร์)", "สมุทรสาคร"),
}


def read_csv_rows(path):
    raw = Path(path).read_bytes()
    last_error = None
    for encoding in ("utf-8-sig", "cp874", "tis-620"):
        try:
            text = raw.decode(encoding)
            return list(csv.reader(text.splitlines()))
        except UnicodeDecodeError as error:
            last_error = error
    raise ValueError(f"อ่านภาษาไทยในไฟล์ CSV ไม่สำเร็จ: {last_error}")


def cell(row, index):
    if index >= len(row):
        return ""
    return str(row[index] or "").strip()


def parse_number(value):
    cleaned = str(value or "").replace(",", "").strip()
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def normalize_warehouse(value):
    compact = re.sub(r"[\s\d()（）._-]+", "", str(value or ""))
    if "นครราชสีมา" in compact or "โคราช" in compact:
        return "โคราช"
    if "สุราษฎร์" in compact:
        return "สุราษฎร์ธานี"
    for warehouse in WAREHOUSE_RULES:
        if warehouse in compact:
            return warehouse
    return re.sub(r"\s*\d+\s*$", "", str(value or "").strip())


def is_valueplus_customer(value):
    compact = re.sub(r"\s+", "", str(value or ""))
    return "แวลูพลัส" in compact and ("เวิร์ลไวด์" in compact or "เวิลด์ไวด์" in compact)


def is_cpall_customer(value):
    compact = re.sub(r"\s+", "", str(value or ""))
    return "ซีพีออลล์" in compact


def resolve_customer(source_customer, warehouse):
    if is_valueplus_customer(source_customer):
        return "Value Plus World Wide", "ปทุมธานี", "ready", ""
    if is_cpall_customer(source_customer):
        rule = WAREHOUSE_RULES.get(warehouse)
        if rule:
            return rule[0], rule[1], "ready", ""
        return "", "", "review", f"ไม่รู้จักคลัง: {warehouse or '-'}"
    return "", "", "review", f"ไม่รู้จักลูกค้า: {source_customer or '-'}"


def build_contexts(rows):
    warehouse = ""
    customer = ""
    contexts = {}
    for index, row in enumerate(rows):
        column_b = cell(row, 1)
        column_c = cell(row, 2)
        column_d = cell(row, 3)
        if "เขตการขาย" in column_b and column_c:
            warehouse = normalize_warehouse(column_c)
        for candidate in (column_d, column_c):
            if is_cpall_customer(candidate) or is_valueplus_customer(candidate):
                customer = candidate
        contexts[index] = (warehouse, customer)
    return contexts


def parse_records(csv_path, template_url):
    rows = read_csv_rows(csv_path)
    records = []
    anchors = []
    contexts = build_contexts(rows)
    for index, row in enumerate(rows):
        match = IV_PATTERN.match(cell(row, 6))
        if match:
            anchors.append((index, match))

    for position, (index, match) in enumerate(anchors):
        next_anchor = anchors[position + 1][0] if position + 1 < len(anchors) else len(rows)
        row = rows[index]
        source_invoice = cell(row, 6)
        invoice = match.group(1).upper()
        date = cell(row, 7)
        warehouse, source_customer = contexts[index]
        exc_vat = parse_number(cell(row, 11))
        quantity = 0.0
        item_count = 0
        expected_line = 1

        for item_index in range(index + 1, next_anchor):
            item_row = rows[item_index]
            line_number = cell(item_row, 8)
            item_name = cell(item_row, 9)
            item_quantity = parse_number(cell(item_row, 11))
            valid_item = (
                line_number.isdigit()
                and int(line_number) == expected_line
                and item_name
                and item_quantity is not None
            )
            if valid_item:
                quantity += item_quantity
                item_count += 1
                expected_line += 1
            elif item_count:
                break

        customer, destination, status, message = resolve_customer(source_customer, warehouse)
        issues = []
        if not DATE_PATTERN.match(date):
            issues.append("วันที่ไม่ถูกต้อง")
        if exc_vat is None:
            issues.append("ไม่พบราคา Exc-vat")
        if item_count == 0:
            issues.append("ไม่พบรายการสินค้า")
        if issues:
            status = "error"
            message = ", ".join(issues)

        records.append({
            "source_row": index + 1,
            "date": date,
            "invoice": invoice,
            "source_invoice": source_invoice,
            "source_customer": source_customer,
            "warehouse": warehouse,
            "customer": customer,
            "destination": destination,
            "quantity": int(quantity) if quantity.is_integer() else quantity,
            "exc_vat": exc_vat or 0,
            "item_count": item_count,
            "status": status,
            "message": message,
        })

    if not records:
        raise ValueError("ไม่พบเลข IV ในคอลัมน์ G ของไฟล์ CSV")

    records.sort(
        key=lambda record: (
            int("".join(re.findall(r"\d+", record["invoice"])) or 0),
            record["invoice"],
        ),
    )

    return {
        "csv_path": str(Path(csv_path)),
        "template_url": template_url,
        "record_count": len(records),
        "total_quantity": sum(record["quantity"] for record in records),
        "total_exc_vat": round(sum(record["exc_vat"] for record in records), 2),
        "review_count": sum(record["status"] == "review" for record in records),
        "error_count": sum(record["status"] == "error" for record in records),
        "warehouses": sorted({record["warehouse"] for record in records if record["warehouse"]}),
        "records": records,
        "output_path": "",
    }


def download_template(template_url, destination):
    request = urllib.request.Request(template_url, headers={"User-Agent": "ValuePlus-System/1.0"})
    with urllib.request.urlopen(request, timeout=60) as response:
        Path(destination).write_bytes(response.read())


def export_workbook(result, template_url, output_path):
    from openpyxl import load_workbook

    if result["review_count"] or result["error_count"]:
        raise ValueError("ยังมีรายการที่ต้องตรวจสอบ กรุณาแก้ไขข้อมูลก่อนบันทึก")

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="valueplus-receivables-") as folder:
        template_path = Path(folder) / "template.xlsx"
        download_template(template_url, template_path)
        workbook = load_workbook(template_path)
        if "ลูกหนี้" not in workbook.sheetnames:
            raise ValueError("ไม่พบชีต 'ลูกหนี้' ใน Google Sheet Template")
        sheet = workbook["ลูกหนี้"]
        for offset, record in enumerate(result["records"]):
            row_number = 4 + offset
            values = (
                record["date"],
                record["invoice"],
                record["customer"],
                record["destination"],
                record["quantity"],
                record["exc_vat"],
            )
            for column, value in enumerate(values, start=1):
                sheet.cell(row=row_number, column=column, value=value)
        workbook.save(output)
    result["output_path"] = str(output)
    return result


def main(argv=None):
    parser = argparse.ArgumentParser()
    parser.add_argument("--csv", required=True)
    parser.add_argument("--template-url", required=True)
    parser.add_argument("--output")
    parser.add_argument("--preview", action="store_true")
    args = parser.parse_args(argv)
    result = parse_records(args.csv, args.template_url)
    if not args.preview:
        if not args.output:
            raise ValueError("กรุณาระบุไฟล์ Excel ผลลัพธ์")
        result = export_workbook(result, args.template_url, args.output)
    print(json.dumps({"success": True, "data": result}, ensure_ascii=True))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        print(json.dumps({"success": False, "message": str(error)}, ensure_ascii=True))
        raise SystemExit(1)
