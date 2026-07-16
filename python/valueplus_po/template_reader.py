from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from .constants import (
    PRODUCT_FIRST_ROW,
    PRODUCT_LAST_ROW,
)
from .normalizers import normalize_product_name


@dataclass(frozen=True)
class TemplateProduct:
    name: str
    normalized_name: str
    row: int


@dataclass
class TemplateCatalog:
    sheet_names: set[str]
    data_products: list[TemplateProduct]
    sheet_products: dict[str, list[TemplateProduct]]


class TemplateReadError(ValueError):
    pass


def read_template(
    template_path: str | Path,
) -> TemplateCatalog:
    path = Path(template_path)

    if not path.is_file():
        raise TemplateReadError(
            f"ไม่พบไฟล์ Excel Template: {path}",
        )

    workbook = load_workbook(
        path,
        read_only=True,
        data_only=False,
    )

    try:
        data_sheet_name = find_data_sheet_name(
            workbook.sheetnames,
        )

        if not data_sheet_name:
            raise TemplateReadError(
                "ไม่พบชีต Data ใน Excel Template",
            )

        data_sheet = workbook[data_sheet_name]

        data_products = read_data_products(
            data_sheet,
        )

        if not data_products:
            raise TemplateReadError(
                "ไม่พบรายการสินค้าในชีต Data คอลัมน์ E",
            )

        sheet_products: dict[
            str,
            list[TemplateProduct],
        ] = {}

        for sheet_name in workbook.sheetnames:
            if sheet_name == data_sheet_name:
                continue

            sheet = workbook[sheet_name]
            products: list[TemplateProduct] = []

            for row_number in range(
                PRODUCT_FIRST_ROW,
                PRODUCT_LAST_ROW + 1,
            ):
                name = sheet.cell(
                    row=row_number,
                    column=2,
                ).value

                product = create_product(
                    name=name,
                    row_number=row_number,
                )

                if product:
                    products.append(product)

            sheet_products[sheet_name] = products

        return TemplateCatalog(
            sheet_names=set(workbook.sheetnames),
            data_products=data_products,
            sheet_products=sheet_products,
        )

    finally:
        workbook.close()


def find_data_sheet_name(
    sheet_names: list[str],
) -> str | None:
    for sheet_name in sheet_names:
        if sheet_name.strip().lower() == "data":
            return sheet_name

    return None


def read_data_products(
    data_sheet,
) -> list[TemplateProduct]:
    products: list[TemplateProduct] = []

    # บังคับให้ OpenPyXL คำนวณขนาดชีต
    # สำหรับ Template ที่ไม่มีค่า dimension
    try:
        data_sheet.calculate_dimension(
            force=True,
        )
    except Exception:
        pass

    maximum_row = data_sheet.max_row or 1

    for row_number in range(
        2,
        maximum_row + 1,
    ):
        name = data_sheet.cell(
            row=row_number,
            column=5,
        ).value

        product = create_product(
            name=name,
            row_number=row_number,
        )

        if product:
            products.append(product)

    return products


def create_product(
    name,
    row_number: int,
) -> TemplateProduct | None:
    if name is None:
        return None

    clean_name = str(name).strip()

    if not clean_name:
        return None

    normalized_name = normalize_product_name(
        clean_name,
    )

    if not normalized_name:
        return None

    return TemplateProduct(
        name=clean_name,
        normalized_name=normalized_name,
        row=row_number,
    )