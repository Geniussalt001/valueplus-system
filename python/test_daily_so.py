import tempfile
import unittest

from pathlib import Path

from openpyxl import Workbook, load_workbook

from valueplus_so.so_processor import (
    _write_group_workbook,
)


class DailySoWorkbookTest(unittest.TestCase):
    def test_writes_group_code_as_so_number(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            template_path = root / "template.xlsx"
            output_path = root / "output.xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet.append([
                "SONumber",
                "DueDate",
                "CustomerCode",
                "SupplierName",
                "ItemCode",
                "ItemName",
                "Qty",
                "Price",
                "Status",
            ])
            sheet.append([
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ])
            workbook.create_sheet("data")
            workbook.save(template_path)

            _write_group_workbook(
                template_path=template_path,
                output_path=output_path,
                document_date="31/07/2026",
                group={
                    "code": "Q19",
                    "po_text": "รวม 12 PO",
                    "so_text": (
                        "Q19 รวม 12 PO 31/7/26"
                    ),
                    "po_count": 12,
                    "records": [
                        {
                            "item_code": "01-0000-29",
                            "quantity": 10,
                            "price": 10.3,
                        },
                    ],
                },
            )

            result = load_workbook(
                output_path,
                data_only=False,
            )

            try:
                self.assertEqual(
                    result["Sheet1"]["A2"].value,
                    "Q19 รวม 12 PO 31/7/26",
                )
            finally:
                result.close()


if __name__ == "__main__":
    unittest.main()
