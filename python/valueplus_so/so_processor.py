import re
import shutil
import unicodedata

from collections import OrderedDict
from copy import copy
from dataclasses import dataclass, field
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path

import pdfplumber

from openpyxl import load_workbook


PO_PATTERN = re.compile(
    r"เลขที่\s*:\s*([A-Z]\d+)",
)

DATE_PATTERN = re.compile(
    r"วันที่\s*:\s*(\d{2}/\d{2}/\d{4})",
)

WAREHOUSE_PATTERN = re.compile(
    r"(?:คลัง|ศูนย์กระจายสินค้า)\s+BDC\s+(.+?)"
    r"(?:\s+คลังดี|\s+อ้างถึง|\n)",
)

ITEM_PATTERN = re.compile(
    r"^\s*(\d+)\s+"
    r"(\d{13})\s+"
    r"(.+?)\s+"
    r"1\s+"
    r"([\d,]+\.(?:\d{2})?)\s+"
    r"0\s+"
    r"([\d,]+\.\d{2})\s+",
    re.MULTILINE,
)

Q19_WAREHOUSES = {
    "สำโรง",
    "รังสิต",
    "เชียงใหม่",
    "นครสวรรค์",
    "ขอนแก่น",
    "โคราช",
}

Q20_WAREHOUSES = {
    "มหาชัย",
    "ร่มเกล้า",
    "ชลบุรี",
    "โชคชัย",
    "หาดใหญ่",
    "สุราษฎร์ธานี",
}

BARCODE_ITEM_CODES = {
    "8859898700156": "01-0000-29",
    "8859389704540": "01-0000-10",
    "8859389704717": "01-0000-14",
    "8859389704427": "01-0000-16",
    "8859389704618": "01-0000-18",
    "8859389704731": "01-0000-22",
    "8859389704670": "01-0000-23",
    "8859389704632": "01-0000-24",
    "8859389704656": "01-0000-28",
    "8859898700071": "01-0000-27",
    "8859898700057": "01-0000-26",
    "8859898700118": "01-0000-30",
    "8859898700194": "01-0000-31",
    "8859898700132": "01-0000-32",
    "8859898700255": "01-0000-33",
    "8859898700279": "01-0000-35",
    "8859898700217": "01-0000-34",
    "8859898700316": "01-0000-37",
    "8859898700293": "01-0000-36",
}


class DailySoError(ValueError):
    pass


@dataclass
class PdfItem:
    barcode: str
    pdf_name: str
    quantity: float
    price: float
    page_number: int


@dataclass
class PdfDocument:
    po_number: str
    document_date: str
    warehouse: str
    pages: list[int] = field(
        default_factory=list,
    )
    items: list[PdfItem] = field(
        default_factory=list,
    )


@dataclass
class TemplateProduct:
    item_code: str
    item_name: str
    price: float | None
    row_number: int
    normalized_name: str


def preview_daily_so(
    pdf_path: str | Path,
    template_path: str | Path,
) -> dict:
    source_pdf = _validate_file(
        pdf_path,
        ".pdf",
        "ไม่พบไฟล์ PDF",
    )

    source_template = _validate_file(
        template_path,
        ".xlsx",
        "ไม่พบไฟล์ Data-SO.Import.xlsx",
    )

    documents = _parse_pdf(
        source_pdf,
    )

    template_products = _read_template_products(
        source_template,
    )

    return _build_preview(
        source_pdf,
        source_template,
        documents,
        template_products,
    )


def process_daily_so(
    pdf_path: str | Path,
    template_path: str | Path,
    output_folder: str | Path,
    quantity_overrides: dict[str, float] | None = None,
) -> dict:
    preview = preview_daily_so(
        pdf_path,
        template_path,
    )

    if preview["error_count"] > 0:
        raise DailySoError(
            "ยังไม่สามารถสร้างไฟล์ได้ "
            "เนื่องจากมีรายการที่ต้องตรวจสอบใน Preview",
        )

    destination = Path(
        output_folder,
    ).resolve()

    destination.mkdir(
        parents=True,
        exist_ok=True,
    )

    output_paths = []

    _apply_quantity_overrides(
        preview["groups"],
        quantity_overrides or {},
    )

    for group in preview["groups"]:
        output_path = (
            destination
            / group["output_name"]
        )

        _write_group_workbook(
            template_path=Path(
                preview["template_path"],
            ),
            output_path=output_path,
            document_date=(
                preview["document_date"]
            ),
            group=group,
        )

        output_paths.append(
            str(output_path),
        )

    preview["output_folder"] = str(
        destination,
    )
    preview["output_paths"] = (
        output_paths
    )

    return preview


def _apply_quantity_overrides(
    groups: list[dict],
    overrides: dict[str, float],
) -> None:
    for group in groups:
        for record in group["records"]:
            key = (
                f'{group["code"]}|'
                f'{record["item_code"]}|'
                f'{record["price"]}'
            )

            legacy_key = (
                f'{group["code"]}|'
                f'{record["item_code"]}'
            )

            override_key = (
                key
                if key in overrides
                else legacy_key
            )

            if override_key not in overrides:
                continue

            try:
                next_quantity = float(
                    overrides[override_key],
                )
            except (
                TypeError,
                ValueError,
            ) as error:
                raise DailySoError(
                    f"ยอดที่แก้ไขไม่ถูกต้อง: {key}",
                ) from error

            if next_quantity < 0:
                raise DailySoError(
                    f"ยอดที่แก้ไขต้องไม่น้อยกว่า 0: {key}",
                )

            original_quantity = (
                record["quantity"]
            )
            record["original_quantity"] = (
                original_quantity
            )
            record["quantity"] = (
                _clean_number(
                    next_quantity,
                )
            )
            record["adjusted"] = True

        group["total_quantity"] = (
            _clean_number(
                sum(
                    record["quantity"]
                    for record in group["records"]
                ),
            )
        )


def _validate_file(
    value: str | Path,
    suffix: str,
    message: str,
) -> Path:
    path = Path(
        value,
    ).resolve()

    if not path.is_file():
        raise DailySoError(
            f"{message}: {path}",
        )

    if path.suffix.lower() != suffix:
        raise DailySoError(
            f"ไฟล์ไม่ใช่ชนิด {suffix}: {path}",
        )

    return path


def _parse_pdf(
    pdf_path: Path,
) -> list[PdfDocument]:
    po_map: OrderedDict[
        str,
        PdfDocument,
    ] = OrderedDict()

    with pdfplumber.open(
        pdf_path,
    ) as pdf:
        for page_number, page in enumerate(
            pdf.pages,
            start=1,
        ):
            text = (
                page.extract_text(
                    layout=True,
                )
                or ""
            )

            po_match = PO_PATTERN.search(
                text,
            )

            if not po_match:
                continue

            po_number = po_match.group(
                1,
            ).strip()

            date_match = DATE_PATTERN.search(
                text,
            )

            warehouse_match = (
                WAREHOUSE_PATTERN.search(
                    text,
                )
            )

            if po_number not in po_map:
                po_map[po_number] = (
                    PdfDocument(
                        po_number=po_number,
                        document_date=(
                            date_match.group(1)
                            if date_match
                            else ""
                        ),
                        warehouse=(
                            _normalize_warehouse(
                                warehouse_match.group(1)
                                if warehouse_match
                                else ""
                            )
                        ),
                    )
                )

            document = po_map[
                po_number
            ]

            document.pages.append(
                page_number,
            )

            if (
                not document.document_date
                and date_match
            ):
                document.document_date = (
                    date_match.group(1)
                )

            if (
                not document.warehouse
                and warehouse_match
            ):
                document.warehouse = (
                    _normalize_warehouse(
                        warehouse_match.group(1),
                    )
                )

            for match in ITEM_PATTERN.finditer(
                text,
            ):
                document.items.append(
                    PdfItem(
                        barcode=match.group(
                            2,
                        ),
                        pdf_name=" ".join(
                            match.group(3).split(),
                        ),
                        quantity=_parse_number(
                            match.group(4),
                        ),
                        price=_parse_number(
                            match.group(5),
                        ),
                        page_number=(
                            page_number
                        ),
                    ),
                )

    documents = list(
        po_map.values(),
    )

    if not documents:
        raise DailySoError(
            "ไม่พบเลข PO ในไฟล์ PDF",
        )

    documents.sort(
        key=lambda document: (
            document.po_number
        ),
    )

    return documents


def _read_template_products(
    template_path: Path,
) -> list[TemplateProduct]:
    workbook = load_workbook(
        template_path,
        data_only=False,
        read_only=True,
    )

    try:
        if "Sheet1" not in workbook.sheetnames:
            raise DailySoError(
                "Template ไม่มีชีต Sheet1",
            )

        if "data" not in workbook.sheetnames:
            raise DailySoError(
                "Template ไม่มีชีต data",
            )

        data_sheet = workbook[
            "data"
        ]

        products = []

        for row_number in range(
            2,
            data_sheet.max_row + 1,
        ):
            item_code = str(
                data_sheet.cell(
                    row_number,
                    5,
                ).value
                or ""
            ).strip()

            item_name = str(
                data_sheet.cell(
                    row_number,
                    6,
                ).value
                or ""
            ).strip()

            if not item_code or not item_name:
                continue

            raw_price = data_sheet.cell(
                row_number,
                8,
            ).value

            price = (
                float(raw_price)
                if isinstance(
                    raw_price,
                    (int, float),
                )
                else None
            )

            products.append(
                TemplateProduct(
                    item_code=item_code,
                    item_name=item_name,
                    price=price,
                    row_number=row_number,
                    normalized_name=(
                        _normalize_product_name(
                            item_name,
                        )
                    ),
                ),
            )

        if not products:
            raise DailySoError(
                "ไม่พบรายการสินค้าในชีต data",
            )

        return products
    finally:
        workbook.close()


def _build_preview(
    pdf_path: Path,
    template_path: Path,
    documents: list[PdfDocument],
    template_products: list[TemplateProduct],
) -> dict:
    dates = {
        document.document_date
        for document in documents
        if document.document_date
    }

    if len(dates) != 1:
        raise DailySoError(
            "วันที่ใน PDF ต้องเป็นวันเดียวกันทั้งหมด",
        )

    document_date = next(
        iter(dates),
        "",
    )

    if not document_date:
        raise DailySoError(
            "ไม่พบวันที่เอกสารใน PDF",
        )

    parsed_date = _parse_document_date(
        document_date,
    )

    groups = []
    total_errors = 0

    for group_code in (
        "Q19",
        "Q20",
    ):
        group_documents = [
            document
            for document in documents
            if _warehouse_group(
                document.warehouse,
            ) == group_code
        ]

        group_result = _build_group(
            group_code,
            group_documents,
            template_products,
            parsed_date,
        )

        total_errors += (
            group_result["error_count"]
        )

        groups.append(
            group_result,
        )

    unknown_warehouses = sorted({
        document.warehouse
        for document in documents
        if _warehouse_group(
            document.warehouse,
        ) is None
    })

    if unknown_warehouses:
        total_errors += len(
            unknown_warehouses,
        )

    return {
        "pdf_path": str(
            pdf_path,
        ),
        "template_path": str(
            template_path,
        ),
        "document_date": (
            document_date
        ),
        "output_date": (
            parsed_date.strftime(
                "%d.%m.%Y",
            )
        ),
        "po_count": len(
            documents,
        ),
        "item_line_count": sum(
            len(document.items)
            for document in documents
        ),
        "error_count": (
            total_errors
        ),
        "unknown_warehouses": (
            unknown_warehouses
        ),
        "groups": groups,
        "output_folder": "",
        "output_paths": [],
    }


def _build_group(
    group_code: str,
    documents: list[PdfDocument],
    template_products: list[TemplateProduct],
    parsed_date: datetime,
) -> dict:
    products_by_code = {
        product.item_code: product
        for product in template_products
    }

    records: OrderedDict[
        str,
        dict,
    ] = OrderedDict()

    errors = []

    for document in documents:
        if not document.items:
            errors.append(
                f"{document.po_number} ไม่มีรายการสินค้า ระบบข้าม PO นี้",
            )
            continue

        for item in document.items:
            product, score, method = (
                _match_product(
                    item,
                    template_products,
                    products_by_code,
                )
            )

            if product is None:
                key = (
                    f"UNMATCHED:{item.barcode}:"
                    f"{item.pdf_name}:"
                    f"{_clean_number(item.price)}"
                )

                record = records.get(
                    key,
                )

                if record is None:
                    record = {
                        "item_code": "",
                        "item_name": "",
                        "pdf_name": item.pdf_name,
                        "barcodes": [
                            item.barcode,
                        ],
                        "quantity": 0,
                        "price": item.price,
                        "match_score": score,
                        "match_method": method,
                        "status": "error",
                        "message": (
                            "ไม่พบสินค้าที่ตรงกันในชีต data"
                        ),
                    }

                    records[key] = (
                        record
                    )

                record["quantity"] += (
                    item.quantity
                )
                continue

            key = (
                f"{product.item_code}|"
                f"{_clean_number(item.price)}"
            )
            record = records.get(
                key,
            )

            if record is None:
                record = {
                    "item_code": (
                        product.item_code
                    ),
                    "item_name": (
                        product.item_name
                    ),
                    "pdf_name": (
                        item.pdf_name
                    ),
                    "barcodes": [
                        item.barcode,
                    ],
                    "quantity": 0,
                    "price": item.price,
                    "match_score": score,
                    "match_method": method,
                    "status": "ready",
                    "message": "",
                }

                records[key] = record

            record["quantity"] += (
                item.quantity
            )

            if (
                item.barcode
                not in record["barcodes"]
            ):
                record["barcodes"].append(
                    item.barcode,
                )

            if (
                product.price is not None
                and not _same_price(
                    product.price,
                    item.price,
                )
                and record["status"]
                == "ready"
            ):
                record["status"] = (
                    "review"
                )
                record["message"] = (
                    "ราคา PDF ไม่ตรงกับราคาในชีต data"
                )

    clean_records = []

    for record in records.values():
        record["quantity"] = (
            _clean_number(
                record["quantity"],
            )
        )
        record["price"] = (
            _clean_number(
                record["price"],
            )
        )
        clean_records.append(
            record,
        )

    error_count = sum(
        1
        for record in clean_records
        if record["status"] == "error"
    )

    review_count = sum(
        1
        for record in clean_records
        if record["status"] == "review"
    )

    po_numbers = [
        document.po_number
        for document in documents
    ]

    so_text = _build_so_text(
        group_code,
        len(po_numbers),
        parsed_date,
    )

    return {
        "code": group_code,
        "po_numbers": po_numbers,
        "po_text": (
            f"รวม {len(po_numbers)} PO"
        ),
        "so_text": so_text,
        "warehouses": list(
            OrderedDict.fromkeys(
                document.warehouse
                for document in documents
            ),
        ),
        "po_count": len(
            documents,
        ),
        "item_count": len(
            clean_records,
        ),
        "ready_count": sum(
            1
            for record in clean_records
            if record["status"] == "ready"
        ),
        "review_count": (
            review_count
        ),
        "error_count": (
            error_count
        ),
        "messages": errors,
        "total_quantity": (
            _clean_number(
                sum(
                    record["quantity"]
                    for record in clean_records
                ),
            )
        ),
        "output_name": (
            f"{group_code}-SO.Import "
            f"{parsed_date.strftime('%d.%m.%Y')}.xlsx"
        ),
        "records": clean_records,
    }


def _match_product(
    item: PdfItem,
    products: list[TemplateProduct],
    products_by_code: dict[
        str,
        TemplateProduct,
    ],
) -> tuple[
    TemplateProduct | None,
    float,
    str,
]:
    mapped_code = BARCODE_ITEM_CODES.get(
        item.barcode,
    )

    if mapped_code:
        mapped_product = (
            products_by_code.get(
                mapped_code,
            )
        )

        if mapped_product:
            return (
                mapped_product,
                1.0,
                "barcode",
            )

    normalized_pdf_name = (
        _normalize_product_name(
            item.pdf_name,
        )
    )

    candidates = sorted(
        (
            (
                SequenceMatcher(
                    None,
                    normalized_pdf_name,
                    product.normalized_name,
                ).ratio(),
                product,
            )
            for product in products
        ),
        key=lambda value: value[0],
        reverse=True,
    )

    if not candidates:
        return None, 0.0, "name"

    best_score, best_product = (
        candidates[0]
    )

    second_score = (
        candidates[1][0]
        if len(candidates) > 1
        else 0.0
    )

    if (
        best_score < 0.72
        or best_score - second_score
        < 0.05
    ):
        return (
            None,
            round(
                best_score,
                4,
            ),
            "name",
        )

    return (
        best_product,
        round(
            best_score,
            4,
        ),
        "name",
    )


def _write_group_workbook(
    template_path: Path,
    output_path: Path,
    document_date: str,
    group: dict,
) -> None:
    temporary_path = output_path.with_suffix(
        ".tmp.xlsx",
    )

    shutil.copy2(
        template_path,
        temporary_path,
    )

    workbook = None

    try:
        workbook = load_workbook(
            temporary_path,
        )

        sheet = workbook[
            "Sheet1"
        ]

        records = [
            record
            for record in group[
                "records"
            ]
            if float(
                record["quantity"]
            ) > 0
        ]

        last_row = max(
            sheet.max_row,
            len(records) + 1,
        )

        for row_number in range(
            2,
            last_row + 1,
        ):
            _copy_row_style(
                sheet,
                2,
                row_number,
            )

            for column in range(
                1,
                10,
            ):
                sheet.cell(
                    row_number,
                    column,
                ).value = None

        due_date = _parse_document_date(
            document_date,
        )

        for index, record in enumerate(
            records,
            start=2,
        ):
            sheet.cell(index, 1).value = (
                group.get("so_text")
                or _build_so_text(
                    group["code"],
                    int(group["po_count"]),
                    due_date,
                )
            )

            date_cell = sheet.cell(
                index,
                2,
            )
            date_cell.value = due_date
            date_cell.number_format = (
                "yyyy/m/d"
            )

            sheet.cell(index, 3).value = (
                "MT001"
            )

            sheet.cell(index, 4).value = (
                f'=IF(ISBLANK(C{index}),"",VLOOKUP(C{index},data!C:D,2,0))'
            )

            sheet.cell(index, 5).value = (
                record["item_code"]
            )

            sheet.cell(index, 6).value = (
                f'=IF(ISBLANK(E{index}),"",VLOOKUP(E{index},data!E:F,2,0))'
            )

            sheet.cell(index, 7).value = (
                record["quantity"]
            )

            sheet.cell(index, 8).value = (
                record["price"]
            )

            sheet.cell(index, 9).value = (
                "S"
            )

        workbook.calculation.calcMode = (
            "auto"
        )
        workbook.calculation.fullCalcOnLoad = (
            True
        )
        workbook.calculation.forceFullCalc = (
            True
        )

        workbook.save(
            temporary_path,
        )

        workbook.close()
        workbook = None

        try:
            if output_path.exists():
                output_path.unlink()

            temporary_path.replace(
                output_path,
            )
        except PermissionError as error:
            raise DailySoError(
                "ไม่สามารถบันทึกไฟล์ SO ได้ "
                "กรุณาปิดไฟล์ Q19/Q20 ที่เปิดอยู่ใน Excel "
                f"แล้วลองใหม่อีกครั้ง: {output_path}",
            ) from error
    except Exception:
        temporary_path.unlink(
            missing_ok=True,
        )
        raise
    finally:
        if workbook is not None:
            workbook.close()


def _copy_row_style(
    sheet,
    source_row: int,
    target_row: int,
) -> None:
    sheet.row_dimensions[
        target_row
    ].height = sheet.row_dimensions[
        source_row
    ].height

    for column in range(
        1,
        10,
    ):
        source = sheet.cell(
            source_row,
            column,
        )

        target = sheet.cell(
            target_row,
            column,
        )

        if source.has_style:
            target._style = copy(
                source._style,
            )

        if source.number_format:
            target.number_format = (
                source.number_format
            )

        target.alignment = copy(
            source.alignment,
        )


def _normalize_warehouse(
    value: str,
) -> str:
    normalized = unicodedata.normalize(
        "NFC",
        str(value or ""),
    )

    normalized = normalized.replace(
        "สําโรง",
        "สำโรง",
    )

    normalized = normalized.replace(
        "_คลังดี",
        "",
    )

    normalized = normalized.replace(
        "คลังดี",
        "",
    )

    normalized = " ".join(
        normalized.split(),
    ).strip(
        " _-",
    )

    if normalized in {
        "นครราชสีมา",
        "โคราช",
    }:
        return "โคราช"

    if normalized.startswith(
        "เชียงใหม่",
    ):
        return "เชียงใหม่"

    if normalized.startswith(
        "ขอนแก่น",
    ):
        return "ขอนแก่น"

    return normalized


def _warehouse_group(
    warehouse: str,
) -> str | None:
    if warehouse in Q19_WAREHOUSES:
        return "Q19"

    if warehouse in Q20_WAREHOUSES:
        return "Q20"

    return None


def _normalize_product_name(
    value: str,
) -> str:
    normalized = unicodedata.normalize(
        "NFC",
        str(value or "").lower(),
    )

    normalized = re.sub(
        r"^h(?=[ก-๙])",
        "",
        normalized,
    )

    replacements = {
        "ยูมิยูมิ": "",
        "um": "",
        "ชิฟฟ่อน": "ชิฟฟอน",
        "ช็อคโกแลต": "ช็อกโกแลต",
        "โคโคนัทมิลค์": "มะพร้าว",
        "สอดไส้": "",
        "รส": "",
        "กลิ่น": "",
        "สีชมพู": "นมชมพู",
    }

    for old, new in replacements.items():
        normalized = normalized.replace(
            old,
            new,
        )

    normalized = re.sub(
        r"\d+(?:\.\d+)?\s*(?:g|กรัม)?",
        "",
        normalized,
    )

    return re.sub(
        r"[^a-zก-๙]+",
        "",
        normalized,
    )


def _parse_document_date(
    value: str,
) -> datetime:
    try:
        day, month, year = (
            int(part)
            for part in value.split(
                "/",
            )
        )

        if year >= 2400:
            year -= 543

        return datetime(
            year,
            month,
            day,
        )
    except (
        TypeError,
        ValueError,
    ) as error:
        raise DailySoError(
            f"วันที่เอกสารไม่ถูกต้อง: {value}",
        ) from error


def _build_so_text(
    group_code: str,
    po_count: int,
    document_date: datetime,
) -> str:
    return (
        f"{group_code} รวม {po_count} PO "
        f"{document_date.day}/"
        f"{document_date.month}/"
        f"{document_date.year % 100:02d}"
    )


def _parse_number(
    value: str,
) -> float:
    normalized = str(value).replace(
        ",",
        "",
    ).rstrip(
        ".",
    )

    return float(
        normalized,
    )


def _clean_number(
    value: float,
) -> int | float:
    rounded = round(
        float(value),
        2,
    )

    if rounded.is_integer():
        return int(
            rounded,
        )

    return rounded


def _same_price(
    left: float,
    right: float,
) -> bool:
    return abs(
        float(left)
        - float(right)
    ) < 0.005
